import os
import random
from datetime import datetime
from sqlalchemy import func
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename
from dotenv import load_dotenv, set_key
import subprocess                # ← добавили
import asyncio # <-- ДОБАВИТЬ
from aiogram import Bot # <-- ДОБАВИТЬ

from models import (
    Base, engine, get_session, Employee, Event, Idea, QuizQuestion,
    RoleOnboarding, Topic, RegCode, ArchivedEmployee, Attendance,
    ArchivedAttendance, ArchivedIdea, Role,
    BotText, OnboardingQuestion, OnboardingStep, EmployeeCustomData, RoleGuide, GroupChat
)

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from functools import wraps

# стало
import os
import io
import html
import random
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# --- НАСТРОЙКА ПРИЛОЖЕНИЯ ---
load_dotenv()
app = Flask(__name__)
from datetime import datetime, date, time
@app.template_filter('fmt_dt')
def fmt_dt(value, fmt='%Y-%m-%d %H:%M'):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        try:
            return value.strftime(fmt)
        except Exception:
            pass
    s = str(value)
    patterns = [
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%dT%H:%M',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d.%m.%Y %H:%M',
        '%d.%m.%Y',
    ]
    for p in patterns:
        try:
            dt = datetime.strptime(s, p)
            return dt.strftime(fmt)
        except Exception:
            continue
    return s

@app.template_filter('fmt_date')
def fmt_date(value, fmt="%d.%m.%Y"):
    if not value:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value  # если формат неизвестный, вернуть как есть
    return value.strftime(fmt)

@app.template_filter('fmt_time')
def fmt_time(value, fmt='%H:%M:%S'):
    if value is None:
        return ''
    return value.strftime(fmt)

app.secret_key = os.getenv("FLASK_SECRET_KEY", "a_very_secret_key_for_flask")

UPLOAD_FOLDER_ONBOARDING = 'uploads/onboarding'
UPLOAD_FOLDER_TOPICS = 'uploads/topics'
app.config['UPLOAD_FOLDER_ONBOARDING'] = UPLOAD_FOLDER_ONBOARDING
app.config['UPLOAD_FOLDER_TOPICS'] = UPLOAD_FOLDER_TOPICS

os.makedirs(UPLOAD_FOLDER_ONBOARDING, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_TOPICS, exist_ok=True)
Base.metadata.create_all(engine)

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
if not ADMIN_USERNAME or not ADMIN_PASSWORD:
    raise RuntimeError("ADMIN_USERNAME/ADMIN_PASSWORD не заданы в .env — задайте перед запуском.")


ONBOARDING_DATA_KEYS = {
    'name': 'Имя (обновляет основной профиль)',
    'birthday': 'Дата рождения (обновляет профиль, формат ДД.ММ.ГГГГ)',
    'contact_info': 'Контактная информация (обновляет профиль)',
    'hobby': 'Хобби (дополнительное поле)',
    'favorite_quote': 'Любимая цитата (дополнительное поле)',
    'tshirt_size': 'Размер футболки (дополнительное поле)'
}

def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapped


@app.before_request
def require_login_for_all():
    # Разрешаем доступ без логина для страниц логина, статики и webhooks/healthchecks при желании
    open_paths = {"/login", "/logout"}  # можно пополнять
    if request.path.startswith("/static"):
        return
    if request.path in open_paths:
        return
    if not session.get("is_admin"):
        return redirect(url_for("login", next=request.path))


def get_text(key: str, default: str = "Текст не найден") -> str:
    """Получает текст для бота из БД по ключу."""
    with get_session() as db:
        text_obj = db.get(BotText, key)
        return text_obj.text if text_obj else default

from aiogram.client.bot import DefaultBotProperties

async def _notify_common_chat_async(text: str):
    token = os.getenv("BOT_TOKEN")
    chat_id = os.getenv("COMMON_CHAT_ID")
    if not token or not chat_id:
        print("[notify_common_chat] BOT_TOKEN/COMMON_CHAT_ID не заданы")  # ← явный лог
        return
    try:
        cid = int(str(chat_id).strip())
    except ValueError:
        print(f"[notify_common_chat] COMMON_CHAT_ID не число: {chat_id!r}")
        return

    tg_bot = Bot(token=token, default=DefaultBotProperties(parse_mode="HTML"))
    try:
        await tg_bot.send_message(chat_id=cid, text=text)
    except Exception as e:
        print(f"[notify_common_chat] Ошибка отправки: {e}")
    finally:
        await tg_bot.session.close()

def notify_common_chat(text: str):
    asyncio.run(_notify_common_chat_async(text))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Вы вошли в админ-панель.", "success")
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        else:
            error = "Неверный логин или пароль."
    return render_template("login.html", error=error)

@app.route("/logout", methods=["POST", "GET"])
def logout():
    session.pop("is_admin", None)
    flash("Вы вышли.", "info")
    return redirect(url_for("login"))


# --- ОСНОВНОЙ МАРШРУТ (ГЛАВНАЯ СТРАНИЦА) ---
@app.route('/')
def index():
    with get_session() as db:
        employees = db.query(Employee).filter_by(is_active=True).order_by(Employee.name).all()
        archived_employees = db.query(ArchivedEmployee).order_by(ArchivedEmployee.dismissal_date.desc()).all()
        events = db.query(Event).order_by(Event.event_date.desc()).all()

        # --- ИСПРАВЛЕНИЕ ЗДЕСЬ ---
        ideas = db.query(Idea, Employee.name).outerjoin(Employee, Idea.employee_id == Employee.id).order_by(
            Idea.submission_date.desc()).all()

        topics = db.query(Topic).order_by(Topic.title).all()
        roles = db.query(Role).order_by(Role.name).all()
        bot_texts = db.query(BotText).order_by(BotText.id).all()

        onboarding_constructor_data = {}
        for role in roles:
            questions = db.query(OnboardingQuestion).filter_by(role=role.name).order_by(
                OnboardingQuestion.order_index).all()
            steps = db.query(OnboardingStep).filter_by(role=role.name).order_by(OnboardingStep.order_index).all()
            onboarding_constructor_data[role.name] = {
                "questions": questions,
                "steps": steps
            }

        role_guides_data = {}
        for role in roles:
            guides = db.query(RoleGuide).filter_by(role=role.name).order_by(RoleGuide.order_index).all()
            role_guides_data[role.name] = guides

        attendance_records = db.query(
            Attendance, Employee.name
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).order_by(
            Attendance.date.desc(), Attendance.arrival_time.desc()
        ).all()



        onboarding_data = {}
        for role in roles:
            onboarding_info = db.query(RoleOnboarding).filter_by(role=role.name).first()
            quiz_questions = db.query(QuizQuestion).filter_by(role=role.name).order_by(QuizQuestion.order_index).all()
            onboarding_data[role.name] = {
                "info": onboarding_info,
                "quizzes": quiz_questions
            }

        # НОВОЕ: список групп, где бот админ
        admin_groups = db.query(GroupChat).order_by(GroupChat.name).all()

        config = {
            "BOT_TOKEN": os.getenv("BOT_TOKEN"),
            "DATABASE_URL": os.getenv("DATABASE_URL"),
            "COMMON_CHAT_ID": os.getenv("COMMON_CHAT_ID"),
            "OFFICE_LAT": os.getenv("OFFICE_LAT"),
            "OFFICE_LON": os.getenv("OFFICE_LON"),
            "OFFICE_RADIUS_METERS": os.getenv("OFFICE_RADIUS_METERS")
        }

    return render_template('index.html', employees=employees, archived_employees=archived_employees,
                           events=events, ideas=ideas, topics=topics,
                           onboarding_data=onboarding_data, roles=roles, config=config, bot_texts=bot_texts,
                           onboarding_constructor_data=onboarding_constructor_data, onboarding_data_keys=ONBOARDING_DATA_KEYS,
                           attendance_records=attendance_records, role_guides_data=role_guides_data,     admin_groups=admin_groups,)


# --- НОВЫЕ РОУТЫ ДЛЯ УПРАВЛЕНИЯ ТЕКСТАМИ ---
@app.route('/texts/update/<string:text_id>', methods=['POST'])
def update_text(text_id):
    with get_session() as db:
        text_obj = db.get(BotText, text_id)
        if text_obj:
            text_obj.text = request.form.get('text', '')
            db.commit()
            flash(f"Текст '{text_id}' успешно обновлен.", "success")
        else:
            flash(f"Текст с ключом '{text_id}' не найден.", "danger")
    return redirect(url_for('index')) # В идеале - редирект на нужную вкладку


# --- НОВЫЕ РОУТЫ ДЛЯ УПРАВЛЕНИЯ ВОПРОСАМИ ОНБОРДИНГА ---
@app.route('/onboarding/question/add/<role>', methods=['POST'])
def add_onboarding_question(role):
    with get_session() as db:
        max_idx = db.query(func.max(OnboardingQuestion.order_index)).filter_by(role=role).scalar() or -1
        new_q = OnboardingQuestion(
            role=role,
            question_text=request.form['question_text'],
            data_key=request.form['data_key'],
            order_index=max_idx + 1,
            is_required='is_required' in request.form
        )
        db.add(new_q)
        db.commit()
        flash("Новый вопрос для онбординга добавлен.", "success")
    return redirect(url_for('index'))

@app.route('/onboarding/question/delete/<int:q_id>', methods=['POST'])
def delete_onboarding_question(q_id):
    with get_session() as db:
        q = db.get(OnboardingQuestion, q_id)
        if q:
            db.delete(q)
            db.commit()
            flash("Вопрос онбординга удален.", "warning")
    return redirect(url_for('index'))

@app.route('/onboarding/question/reorder', methods=['POST'])
def reorder_onboarding_question():
    ordered_ids = request.get_json(silent=True).get('ordered_ids', [])
    with get_session() as session:
        for index, qid in enumerate(ordered_ids):
            q = session.get(OnboardingQuestion, int(qid))
            if q:
                q.order_index = index
        session.commit()
    return jsonify(success=True)


# Файл app.py становится намного проще

@app.route('/onboarding/step/add/<role>', methods=['POST'])
def add_onboarding_step(role):
    with get_session() as db:
        max_idx = db.query(func.max(OnboardingStep.order_index)).filter_by(role=role).scalar() or -1
        new_step = OnboardingStep(
            role=role,
            message_text=request.form.get('message_text'),
            order_index=max_idx + 1
        )

        # Логика обработки файла упрощается
        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            # ВАЖНО: JS должен отправлять корректный file_type
            file_type = request.form.get('file_type', 'document')

            # Имя файла можно генерировать на основе типа
            filename = secure_filename(f"{role.lower()}_step_{file.filename}")
            upload_folder = os.path.join(app.config['UPLOAD_FOLDER_ONBOARDING'], 'steps')
            os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)

            file.save(file_path)  # Просто сохраняем готовый файл

            new_step.file_path = file_path
            new_step.file_type = file_type

        db.add(new_step)
        db.commit()
        flash("Новый шаг знакомства добавлен.", "success")
    return redirect(url_for('index'))

@app.route('/onboarding/step/delete/<int:step_id>', methods=['POST'])
def delete_onboarding_step(step_id):
    with get_session() as db:
        step = db.get(OnboardingStep, step_id)
        if step:
            # Опционально: удалить связанный файл с диска
            if step.file_path and os.path.exists(step.file_path):
                os.remove(step.file_path)
            db.delete(step)
            db.commit()
            flash("Шаг знакомства удален.", "warning")
    return redirect(url_for('index'))

@app.route('/onboarding/step/reorder', methods=['POST'])
def reorder_onboarding_step():
    ordered_ids = request.get_json(silent=True).get('ordered_ids', [])
    with get_session() as session:
        for index, sid in enumerate(ordered_ids):
            step = session.get(OnboardingStep, int(sid))
            if step:
                step.order_index = index
        session.commit()
    return jsonify(success=True)


# --- УПРАВЛЕНИЕ РОЛЯМИ (СПЕЦИАЛЬНОСТЯМИ) ---
@app.route('/role/add', methods=['POST'])
def add_role():
    with get_session() as db:
        role_name = request.form.get('role_name')
        if role_name and not db.query(Role).filter_by(name=role_name).first():
            db.add(Role(name=role_name))
            db.commit()
            flash(f"Роль '{role_name}' успешно добавлена.", "success")
        else:
            flash(f"Роль '{role_name}' уже существует или имя не указано.", "danger")
    return redirect(url_for('index'))


@app.route('/role/delete/<int:role_id>', methods=['POST'])
def delete_role(role_id):
    with get_session() as db:
        role = db.get(Role, role_id)
        if role:
            db.delete(role)
            db.commit()
            flash(f"Роль '{role.name}' удалена.", "warning")
    return redirect(url_for('index'))


# --- УПРАВЛЕНИЕ СОТРУДНИКАМИ ---
@app.route('/employee/add', methods=['POST'])
def add_employee():
    with get_session() as db:
        email = request.form['email']
        role = request.form['role'] # Добавили получение роли
        if db.query(Employee).filter_by(email=email).first():
            flash(f"Сотрудник с email {email} уже существует!", "danger")
            return redirect(url_for('index'))

        # Создаем сотрудника только с email и ролью. Имя = email как заглушка.
        new_emp = Employee(
            email=email,
            name=email, # Используем email как временное имя
            role=role,
            is_active=True
        )
        db.add(new_emp)
        db.flush()

        while True:
            code = "".join(str(random.randint(0, 9)) for _ in range(8))
            if not db.query(RegCode).filter_by(code=code).first(): break

        db.add(RegCode(code=code, email=new_emp.email, used=False))
        db.commit()
        flash(f"Сотрудник с email {email} добавлен. Код для регистрации: {code}", "success")
    return redirect(url_for('index'))

@app.route('/employee/reset_progress/<int:emp_id>', methods=['POST'])
def reset_progress(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            flash("Сотрудник не найден!", "danger")
            return redirect(url_for('index'))

        # Сбрасываем прогресс
        emp.onboarding_completed = False
        emp.training_passed = False

        # Удаляем его кастомные ответы на вопросы онбординга
        db.query(EmployeeCustomData).filter_by(employee_id=emp_id).delete(synchronize_session=False)

        db.commit()

        # Асинхронно отправляем уведомление в Telegram, если есть привязка
        if emp.telegram_id:
            async def send_notification():
                token = os.getenv("BOT_TOKEN")
                # Убедитесь, что в вашей таблице bot_texts есть ключ 'progress_reset_notification'
                message_text = get_text('progress_reset_notification',
                                        'Администратор сбросил ваш прогресс. Вам потребуется заново пройти онбординг и тренинг при следующем входе в бота.')
                if not token:
                    flash("Токен бота не найден в .env, уведомление не отправлено.", "warning")
                    return

                tg_bot = Bot(token=token)
                try:
                    await tg_bot.send_message(emp.telegram_id, message_text)
                except Exception as e:
                    flash(f"Не удалось отправить уведомление: {e}", "danger")
                    print(f"Could not send reset notification to {emp.id}: {e}")
                finally:
                    await tg_bot.session.close()

            try:
                asyncio.run(send_notification())
            except RuntimeError: # Если уже есть запущенный event loop
                 # Это может произойти в некоторых средах. Более сложная обработка может потребоваться.
                 # Для простоты пока просто логируем.
                 print("Could not run asyncio.run(), possibly due to existing loop.")


        flash(f"Прогресс для сотрудника {emp.name or emp.email} полностью сброшен.", "warning")
    return redirect(url_for('index'))


@app.route('/broadcast/send', methods=['POST'])
def send_broadcast():
    message_text = request.form.get('message_text')
    target_role = request.form.get('target_role')

    if not message_text:
        flash("Текст сообщения не может быть пустым.", "danger")
        return redirect(url_for('index'))

    with get_session() as db:
        query = db.query(Employee.telegram_id).filter(Employee.is_active == True, Employee.telegram_id != None)
        if target_role != 'all':
            query = query.filter(Employee.role == target_role)

        target_users_ids = [row[0] for row in query.all()]

    if not target_users_ids:
        flash("Не найдено сотрудников для рассылки в выбранном сегменте.", "warning")
        return redirect(url_for('index'))

    token = os.getenv("BOT_TOKEN")
    if not token:
        flash("Токен бота не найден в .env, рассылка невозможна.", "danger")
        return redirect(url_for('index'))

    # Статистика отправки
    success_count = 0
    error_count = 0

    async def _send_to_all():
        nonlocal success_count, error_count
        tg_bot = Bot(token=token)
        try:
            for user_id in target_users_ids:
                try:
                    await tg_bot.send_message(chat_id=user_id, text=message_text)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    print(f"Failed to send message to {user_id}: {e}")
                await asyncio.sleep(0.1)  # Пауза во избежание лимитов Telegram
        finally:
            await tg_bot.session.close()

    try:
        asyncio.run(_send_to_all())
    except RuntimeError:
        flash("Ошибка выполнения асинхронной задачи рассылки.", "danger")
        return redirect(url_for('index'))

    flash(f"Рассылка завершена. Успешно отправлено: {success_count}. Ошибок: {error_count}.", "success")
    return redirect(url_for('index'))

@app.route('/employee/edit/<int:emp_id>', methods=['POST'])
def edit_employee(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            flash("Сотрудник не найден!", "danger")
            return redirect(url_for('index'))

        new_email = request.form['email']
        if emp.email != new_email and db.query(Employee).filter_by(email=new_email).first():
            flash(f"Email {new_email} уже занят другим сотрудником!", "danger")
            return redirect(url_for('index'))

        old_role = emp.role  # ← запоминаем старую роль

        emp.name = request.form['name']
        emp.email = new_email
        emp.role = request.form['role']
        emp.birthday = datetime.strptime(request.form['birthday'], '%Y-%m-%d').date() if request.form[
            'birthday'] else None
        db.commit()

        # уведомление в общий чат при смене роли
        if emp.role != old_role:
            promo_text = get_text(
                'employee_role_changed_announcement',
                '🎉 <b>{name}</b>: {old} → {new}'
            ).format(
                name=html.escape(emp.name or emp.email),
                old=html.escape(old_role or '—'),
                new=html.escape(emp.role or '—')
            )
            notify_common_chat(promo_text)

        flash(f"Данные сотрудника {emp.name} обновлены.", "success")
    return redirect(url_for('index'))


@app.route('/employee/dismiss/<int:emp_id>', methods=['POST'])
def dismiss_employee(emp_id):
    try:
        with get_session() as db:
            emp = db.get(Employee, emp_id)
            if not emp:
                flash("Сотрудник не найден!", "danger")
                return redirect(url_for('index'))

            name_cache = emp.name or emp.email
            role_cache = emp.role or ''

            emp_data = {
                "original_employee_id": emp.id,
                "telegram_id": emp.telegram_id,
                "email": emp.email,
                "role": emp.role,
                "name": emp.name,
                "birthday": emp.birthday,
                "registered": emp.registered,
                "training_passed": emp.training_passed,
            }
            archived_emp = ArchivedEmployee(**emp_data)
            db.add(archived_emp)

            # 2) Архивируем посещаемость
            for att in db.query(Attendance).filter_by(employee_id=emp_id).all():
                db.add(ArchivedAttendance(
                    employee_id=att.employee_id,
                    date=att.date,
                    arrival_time=att.arrival_time,
                    departure_time=att.departure_time
                ))
                db.delete(att)

            # 3) Архивируем идеи
            for idea in db.query(Idea).filter_by(employee_id=emp_id).all():
                db.add(ArchivedIdea(
                    employee_id=idea.employee_id,
                    idea_text=idea.text,  # <- корректное имя поля
                    submission_date=idea.submission_date
                ))
                db.delete(idea)

            db.delete(emp)
            db.commit()

            dismiss_text = get_text(
                'employee_dismissed_announcement',
                '👋 <b>{name}</b> ({role}) больше не работает с нами. Пожелаем удачи!'
            ).format(name=html.escape(name_cache), role=html.escape(role_cache))

        flash(f"Сотрудник {emp.name} уволен и перенесен в архив.", "warning")
        return redirect(url_for('index'))

    except Exception as e:
        # На всякий случай всегда возвращаем redirect, чтобы не ловить TypeError от Flask
        print(f"[dismiss_employee] error: {e}")
        flash("Ошибка при увольнении сотрудника.", "danger")
        return redirect(url_for('index'))

@app.route('/employee/reset_telegram/<int:emp_id>', methods=['POST'])
def reset_telegram(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            flash("Сотрудник не найден!", "danger")
            return redirect(url_for('index'))
        emp.telegram_id = None
        emp.registered = False
        db.commit()
        flash(f"Привязка Telegram для сотрудника {emp.name} сброшена. Ему потребуется новый код для входа.", "warning")
    return redirect(url_for('index'))


@app.route('/employee/generate_code/<int:emp_id>', methods=['POST'])
def generate_new_code(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            flash("Сотрудник не найден!", "danger")
            return redirect(url_for('index'))
        while True:
            code = "".join(str(random.randint(0, 9)) for _ in range(8))
            if not db.query(RegCode).filter_by(code=code).first(): break
        db.add(RegCode(code=code, email=emp.email, used=False))
        db.commit()
        flash(f"Сгенерирован новый код для {emp.name}: {code}", "success")
    return redirect(url_for('index'))


# --- УПРАВЛЕНИЕ ОНБОРДИНГОМ И КВИЗАМИ ---
@app.route('/onboarding/update/<role>', methods=['POST'])
def update_onboarding(role):
    with get_session() as db:
        onboarding = db.query(RoleOnboarding).filter_by(role=role).first()
        if not onboarding:
            onboarding = RoleOnboarding(role=role)
            db.add(onboarding)

        onboarding.text = request.form['text']
        onboarding.file_type = request.form['file_type']

        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            filename = secure_filename(f"{role.lower()}_{file.filename}")
            file_path = os.path.join(app.config['UPLOAD_FOLDER_ONBOARDING'], filename)
            file.save(file_path)
            onboarding.file_path = file_path

        db.commit()
        flash(f"Онбординг для роли '{role}' обновлен.", "success")
    return redirect(url_for('index'))


@app.route('/quiz/add/<role>', methods=['POST'])
def add_quiz(role):
    with get_session() as db:
        qtype    = request.form['question_type']
        question = request.form['question']

        if qtype == 'choice':
            options = request.form.get('options')        # "Вариант1;Вариант2;…"
            answer  = request.form.get('answer')         # из скрытого поля
        else:
            options = None
            answer  = request.form.get('text_answer')    # из текстового поля

        # вычисляем следующий индекс
        max_idx = db.query(func.max(QuizQuestion.order_index)).filter_by(role=role).scalar() or -1

        new_q = QuizQuestion(
            role=role,
            question=question,
            answer=answer,
            question_type=qtype,
            options=options,
            order_index=max_idx + 1
        )
        db.add(new_q)
        db.commit()
        flash("Новый вопрос добавлен.", "success")
    return redirect(url_for('index'))

@app.route('/quiz/edit/<int:quiz_id>', methods=['POST'])
def edit_quiz(quiz_id):
    with get_session() as db:
        quiz = db.get(QuizQuestion, quiz_id)
        qtype = request.form['question_type']

        quiz.question = request.form['question']
        if qtype == 'choice':
            quiz.options = request.form.get('options')
            quiz.answer  = request.form.get('answer')
        else:
            quiz.options = None
            quiz.answer  = request.form.get('text_answer')

        quiz.question_type = qtype
        db.commit()
        flash("Вопрос обновлён.", "success")
    return redirect(url_for('index'))

@app.route('/quiz/delete/<int:quiz_id>', methods=['POST'])
def delete_quiz(quiz_id):
    with get_session() as db:
        quiz = db.get(QuizQuestion, quiz_id)
        if quiz:
            db.delete(quiz)
            db.commit()
            flash("Вопрос квиза удален.", "warning")
    return redirect(url_for('index'))


@app.route('/quiz/reorder', methods=['POST'])
def reorder_quiz():
    data = request.get_json(silent=True) or {}
    ordered_ids = data.get('ordered_ids', [])
    # используем session вместо db, чтобы не получить NameError
    with get_session() as session:
        for index, qid in enumerate(ordered_ids):
            quiz = session.get(QuizQuestion, int(qid))
            if quiz:
                quiz.order_index = index
        session.commit()
    return jsonify(success=True, message="Порядок вопросов обновлен")

# --- УПРАВЛЕНИЕ ИВЕНТАМИ ---
@app.route('/event/add', methods=['POST'])
def add_event():
    with get_session() as db:
        title = request.form['title']
        description = request.form['description']
        event_dt = datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M')

        db.add(Event(title=title, description=description, event_date=event_dt))
        db.commit()

        # уведомление в общий чат
        evt_text = get_text(
            'event_created_announcement',
            '📅 Новый ивент: <b>{title}</b>\n🗓 {when}\n<i>{desc}</i>'
        ).format(
            title=html.escape(title),
            when=event_dt.strftime('%d.%m.%Y %H:%M'),
            desc=html.escape(description)
        )
        notify_common_chat(evt_text)

        flash("Новый ивент успешно добавлен.", "success")
    return redirect(url_for('index'))



@app.route('/event/edit/<int:event_id>', methods=['POST'])
def edit_event(event_id):
    with get_session() as db:
        event = db.get(Event, event_id)
        if event:
            event.title = request.form['title']
            event.description = request.form['description']
            event.event_date = datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M')
            db.commit()
            flash("Ивент обновлен.", "success")
    return redirect(url_for('index'))


@app.route('/event/delete/<int:event_id>', methods=['POST'])
def delete_event(event_id):
    with get_session() as db:
        event = db.get(Event, event_id)
        if event:
            db.delete(event)
            db.commit()
            flash("Ивент удален.", "warning")
    return redirect(url_for('index'))


# --- УПРАВЛЕНИЕ ИДЕЯМИ ---
@app.route('/idea/delete/<int:idea_id>', methods=['POST'])
def delete_idea(idea_id):
    with get_session() as db:
        idea = db.get(Idea, idea_id)
        if idea:
            db.delete(idea)
            db.commit()
            flash("Идея удалена.", "warning")
    return redirect(url_for('index'))


# --- УПРАВЛЕНИЕ БАЗОЙ ЗНАНИЙ (ТОПИКАМИ) ---
@app.route('/topic/add', methods=['POST'])
def add_topic():
    with get_session() as db:
        new_topic = Topic(title=request.form['title'], content=request.form['content'])
        if 'image' in request.files and request.files['image'].filename != '':
            file = request.files['image']
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER_TOPICS'], filename)
            file.save(file_path)
            new_topic.image_path = file_path
        db.add(new_topic)
        db.commit()
        flash("Новая тема в Базе Знаний создана.", "success")
    return redirect(url_for('index'))


@app.route('/topic/edit/<int:topic_id>', methods=['POST'])
def edit_topic(topic_id):
    with get_session() as db:
        topic = db.get(Topic, topic_id)
        if topic:
            topic.title = request.form['title']
            topic.content = request.form['content']
            if 'image' in request.files and request.files['image'].filename != '':
                file = request.files['image']
                if topic.image_path and os.path.exists(topic.image_path):
                    os.remove(topic.image_path)
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER_TOPICS'], filename)
                file.save(file_path)
                topic.image_path = file_path
            db.commit()
            flash("Тема в Базе Знаний обновлена.", "success")
    return redirect(url_for('index'))


@app.route('/topic/delete/<int:topic_id>', methods=['POST'])
def delete_topic(topic_id):
    with get_session() as db:
        topic = db.get(Topic, topic_id)
        if topic:
            if topic.image_path and os.path.exists(topic.image_path):
                os.remove(topic.image_path)
            db.delete(topic)
            db.commit()
            flash("Тема из Базы Знаний удалена.", "warning")
    return redirect(url_for('index'))


# --- УПРАВЛЕНИЕ НАСТРОЙКАМИ (.ENV) ---
@app.route('/config/update', methods=['POST'])
def update_config():
    env_file = '.env'

    # НОВОЕ: если выбрали из списка — используем его; иначе — поле ручного ввода
    selected = request.form.get("COMMON_CHAT_ID_SELECT", "")
    manual   = request.form.get("COMMON_CHAT_ID", "")

    common_chat_id = selected or manual  # приоритет у выпадающего списка
    set_key(env_file, "COMMON_CHAT_ID", common_chat_id)
    os.environ["COMMON_CHAT_ID"] = common_chat_id  # ← применяем сразу в текущем процессе

    set_key(env_file, "OFFICE_LAT", request.form.get("OFFICE_LAT", ""))
    set_key(env_file, "OFFICE_LON", request.form.get("OFFICE_LON", ""))
    set_key(env_file, "OFFICE_RADIUS_METERS", request.form.get("OFFICE_RADIUS_METERS", ""))
    os.environ["OFFICE_LAT"] = request.form.get("OFFICE_LAT", "")
    os.environ["OFFICE_LON"] = request.form.get("OFFICE_LON", "")
    os.environ["OFFICE_RADIUS_METERS"] = request.form.get("OFFICE_RADIUS_METERS", "")

    if request.form.get("BOT_TOKEN"):
        set_key(env_file, "BOT_TOKEN", request.form.get("BOT_TOKEN"))
        os.environ["BOT_TOKEN"] = request.form.get("BOT_TOKEN")  # ← тоже сразу

    flash(
        "Настройки сохранены. Для отдельного процесса бота всё ещё нужен перезапуск, а панель уже применяет новые значения.",
        "info")
    return redirect(url_for('index'))


@app.route('/guide/add/<role>', methods=['POST'])
def add_guide(role):
    with get_session() as db:
        max_idx = db.query(func.max(RoleGuide.order_index)).filter_by(role=role).scalar() or -1
        new_guide = RoleGuide(
            role=role,
            title=request.form['title'],
            content=request.form.get('content', ''),
            order_index=max_idx + 1
        )

        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            # Папку для регламентов лучше сделать отдельной
            upload_folder = os.path.join(app.config['UPLOAD_FOLDER_ONBOARDING'], 'guides')
            os.makedirs(upload_folder, exist_ok=True)
            filename = secure_filename(f"{role.lower()}_guide_{file.filename}")
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            new_guide.file_path = file_path

        db.add(new_guide)
        db.commit()
        flash(f"Новый регламент для роли '{role}' добавлен.", "success")
    return redirect(url_for('index'))

@app.route('/guide/delete/<int:guide_id>', methods=['POST'])
def delete_guide(guide_id):
    with get_session() as db:
        guide = db.get(RoleGuide, guide_id)
        if guide:
            if guide.file_path and os.path.exists(guide.file_path):
                os.remove(guide.file_path)
            db.delete(guide)
            db.commit()
            flash("Регламент удален.", "warning")
    return redirect(url_for('index'))

# Можно также добавить роуты для редактирования и сортировки по аналогии с другими разделами
# --- -------------------------------------------------------- ---

@app.route('/export/employees.xlsx', methods=['GET'])
def export_employees_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    rows = []
    with get_session() as db:
        employees = db.query(Employee).order_by(Employee.name).all()

        # заранее подтянем неиспользованные коды для существующих сотрудников
        emails = [e.email for e in employees]
        existing_free = {
            rc.email: rc.code
            for rc in db.query(RegCode).filter(
                RegCode.used == False,
                RegCode.email.in_(emails)
            ).all()
        }

        for emp in employees:
            code_for_row = ""
            if not emp.telegram_id:
                # берём уже существующий свободный код, если есть
                code_for_row = existing_free.get(emp.email, "")
                # если нет — генерим уникальный и добавляем (без коммита прямо сейчас)
                if not code_for_row:
                    while True:
                        code = "".join(str(random.randint(0, 9)) for _ in range(8))
                        if not db.query(RegCode).filter_by(code=code).first():
                            break
                    db.add(RegCode(code=code, email=emp.email, used=False))
                    code_for_row = code

            rows.append([
                emp.id,
                emp.name or "",
                emp.email,
                emp.role or "",
                emp.birthday.strftime("%d.%m.%Y") if emp.birthday else "",
                "Да" if emp.is_active else "Нет",
                "Да" if emp.registered else "Нет",
                "Да" if emp.training_passed else "Нет",
                emp.telegram_id or "",
                code_for_row
            ])

        # один коммит в конце — фиксируем новосозданные RegCode
        db.commit()

    # --- оформление XLSX ниже без изменений вашей логики ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    headers = [
        "ID", "ФИО", "Email", "Роль", "Дата рождения",
        "Активен", "Зарегистрирован", "Прошёл тренинг",
        "Telegram ID", "Код для привязки"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c_idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = left if c_idx in (2, 3, 4, 10) else center

    widths = [6, 28, 28, 18, 14, 10, 16, 16, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    ws.conditional_formatting.add(
        f"A2:{get_column_letter(len(headers))}{ws.max_row}",
        FormulaRule(formula=[f'LEN($I2)=0'], stopIfTrue=False,
                    fill=PatternFill("solid", fgColor="FEE2E2"))
    )
    ws.conditional_formatting.add(
        f"J2:J{ws.max_row}",
        FormulaRule(formula=[f'LEN($J2)>0'], stopIfTrue=False,
                    fill=PatternFill("solid", fgColor="ECFDF5"))
    )

    info = wb.create_sheet("Справка")
    info["A1"] = "Экспорт сотрудников"; info["A1"].font = Font(bold=True, size=14)
    info["A3"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    info["A5"] = "Примечание:"
    info["A6"] = "— У сотрудников без Telegram ID сгенерирован 8-значный код привязки (последний столбец)."
    info["A7"] = "— Строки без Telegram ID в основном листе подсвечены красным (для контроля)."

    mem = io.BytesIO()
    wb.save(mem); mem.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"employees_{ts}.xlsx"
    )

@app.route('/landing')
def landing():
    return render_template("landing.html")

# --- ЗАПУСК ПРИЛОЖЕНИЯ ---
if __name__ == '__main__':
app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)), debug=False, use_reloader=False)
